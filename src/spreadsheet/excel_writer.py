"""Excel (.xlsx) への書き出しモジュール"""

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def write_to_excel(data: dict, output_path: str) -> str:
    """集計データをExcelファイルに書き出す。

    Args:
        data: data_processor.process_data() の出力
        output_path: 出力Excelファイルのパス

    Returns:
        出力ファイルの絶対パス
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # ダッシュボード
        _write_dashboard(writer, data["stats"])

        # 月次サマリー
        if not data["monthly_summary"].empty:
            data["monthly_summary"].to_excel(
                writer, sheet_name="月次サマリー", index=False
            )

        # カテゴリ別集計
        if not data["category_breakdown"].empty:
            data["category_breakdown"].to_excel(
                writer, sheet_name="カテゴリ別集計", index=False
            )

        # 月別カテゴリ
        if not data["monthly_category"].empty:
            data["monthly_category"].to_excel(
                writer, sheet_name="月別カテゴリ", index=False
            )

        # 取引一覧
        if not data["transactions"].empty:
            data["transactions"].to_excel(
                writer, sheet_name="取引一覧", index=False
            )

        # 書式設定
        _apply_formatting(writer)

    return str(output_path.resolve())


def _write_dashboard(writer: pd.ExcelWriter, stats: dict) -> None:
    """ダッシュボードシートを作成する。"""
    rows = [
        {"項目": "マネーフォワード 収支レポート", "値": ""},
        {"項目": "", "値": ""},
        {"項目": "集計期間（月数）", "値": stats["num_months"]},
        {"項目": "取引件数", "値": stats["num_transactions"]},
        {"項目": "", "値": ""},
        {"項目": "総収入", "値": f"¥{stats['total_income']:,}"},
        {"項目": "総支出", "値": f"¥{stats['total_expense']:,}"},
        {"項目": "純収支", "値": f"¥{stats['net']:,}"},
        {"項目": "", "値": ""},
        {"項目": "月平均収入", "値": f"¥{stats['avg_monthly_income']:,}"},
        {"項目": "月平均支出", "値": f"¥{stats['avg_monthly_expense']:,}"},
        {"項目": "", "値": ""},
        {"項目": "支出最多カテゴリ", "値": stats["top_expense_category"]},
    ]
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="ダッシュボード", index=False, header=False)


def _apply_formatting(writer: pd.ExcelWriter) -> None:
    """全シートに書式を適用する。"""
    wb = writer.book

    # 共通スタイル
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for ws in wb.worksheets:
        if ws.title == "ダッシュボード":
            _format_dashboard(ws)
            continue

        # ヘッダー行のスタイル
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # データ行のスタイル
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
                # 金額列は右寄せ、カンマ区切り
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "#,##0"

        # 列幅を自動調整
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        # 日本語は2文字分としてカウント
                        val = str(cell.value)
                        length = sum(2 if ord(c) > 127 else 1 for c in val)
                        max_length = max(max_length, length)
            ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

        # ヘッダー行を固定
        ws.freeze_panes = "A2"


def _format_dashboard(ws) -> None:
    """ダッシュボードシートの書式を設定する。"""
    # タイトル
    ws["A1"].font = Font(bold=True, size=16, color="2F5496")

    # セクションヘッダー
    section_font = Font(bold=True, size=12, color="4472C4")
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).startswith("---"):
                cell.font = section_font

    # 列幅
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20

    # B列を右寄せ
    for row in ws.iter_rows(min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="right")
